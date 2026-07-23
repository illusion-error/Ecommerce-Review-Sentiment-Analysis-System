"""FastAPI backend for e-commerce review sentiment analysis."""

from __future__ import annotations

import csv
import io
import uuid
from contextlib import asynccontextmanager
from typing import Any, Dict, List
from urllib.parse import quote

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from .cache import prediction_cache
from .database import (
    analysis_records_for_insights,
    create_batch_task,
    get_batch_task,
    init_db,
    insert_analysis_record,
    list_history,
    records_by_task,
    summary_statistics,
)
from .schemas import SingleSentimentRequest
from .sentiment import predict_sentiment
from .settings import settings
from model.text_insights import extract_keywords, generate_rule_summary, score_aspects


@asynccontextmanager
async def lifespan(_: FastAPI):
    init_db()
    yield


app = FastAPI(title=settings.app_name, version=settings.app_version, lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Make direct TestClient usage and local scripts stable even when lifespan hooks
# are not entered explicitly. `startup()` keeps the same initialization for
# normal uvicorn runs, and SQLite `CREATE TABLE IF NOT EXISTS` is idempotent.
init_db()


def api_response(data: Any = None, message: str = "ok", code: int = 0, success: bool = True) -> Dict[str, Any]:
    return {"success": success, "message": message, "data": data, "code": code}


EXPORT_COLUMNS = [
    ("id", "记录ID"),
    ("task_id", "批量任务ID"),
    ("product_id", "商品ID"),
    ("raw_text", "原始评论"),
    ("clean_text", "清洗后评论"),
    ("label_text", "情感标签"),
    ("sentiment_text", "情感结果"),
    ("confidence", "置信度"),
    ("strength", "情绪强度(0-10)"),
    ("cached_text", "是否命中缓存"),
    ("created_at", "分析时间"),
]


def _format_export_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Convert raw database rows into a readable report table.

    The database keeps machine-friendly fields such as `label=1` and
    `sentiment=positive`. The downloaded report is for people to read in Excel,
    so it uses Chinese headers, stable column order and rounded numeric values.
    """

    formatted: List[Dict[str, Any]] = []
    for row in records:
        sentiment = str(row.get("sentiment") or "")
        try:
            confidence = round(float(row.get("confidence") or 0.0), 4)
        except (TypeError, ValueError):
            confidence = 0.0
        try:
            strength = round(float(row.get("strength") or 0.0), 2)
        except (TypeError, ValueError):
            strength = 0.0
        try:
            label = int(row.get("label") or 0)
        except (TypeError, ValueError):
            label = 0

        item = {
            "id": row.get("id", ""),
            "task_id": row.get("task_id", ""),
            "product_id": row.get("product_id", ""),
            "raw_text": row.get("raw_text", ""),
            "clean_text": row.get("clean_text", ""),
            "label_text": "正向" if label == 1 else "负向",
            "sentiment_text": "正向 positive" if sentiment == "positive" else "负向 negative",
            "confidence": confidence,
            "strength": strength,
            "cached_text": "是" if bool(row.get("cached")) else "否",
            "created_at": row.get("created_at", ""),
        }
        formatted.append({header: item[key] for key, header in EXPORT_COLUMNS})
    return formatted


def _content_disposition(filename: str) -> str:
    quoted = quote(filename)
    return f"attachment; filename={quoted}; filename*=UTF-8''{quoted}"


def _autosize_excel_columns(worksheet, dataframe: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    worksheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    width_limits = {
        "原始评论": 42,
        "清洗后评论": 42,
        "批量任务ID": 22,
        "分析时间": 20,
    }
    min_widths = {
        "原始评论": 24,
        "清洗后评论": 24,
        "情感结果": 14,
    }
    for index, column_name in enumerate(dataframe.columns, start=1):
        values = [str(column_name)] + ["" if pd.isna(value) else str(value) for value in dataframe[column_name].tolist()]
        max_len = max(len(value) for value in values)
        min_width = min_widths.get(str(column_name), 10)
        width = min(max(max_len + 2, min_width), width_limits.get(str(column_name), 18))
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_name in ("置信度", "情绪强度(0-10)"):
        if column_name in dataframe.columns:
            col_idx = dataframe.columns.get_loc(column_name) + 1
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for item in cell:
                    item.number_format = "0.00"


def _analyze_and_store(
    text: str,
    *,
    product_id: str = "",
    task_id: str = "",
    comment_time: str = "",
) -> Dict[str, Any]:
    text = (text or "").strip()
    if not text:
        raise ValueError("评论文本不能为空")

    cached_value = prediction_cache.get(text)
    cached = cached_value is not None
    result = cached_value if cached_value is not None else predict_sentiment(text)
    if not cached:
        prediction_cache.set(text, result)

    stored = insert_analysis_record(
        task_id=task_id,
        product_id=product_id,
        raw_text=text,
        clean_text=result["clean_text"],
        label=int(result["label"]),
        sentiment=str(result["sentiment"]),
        confidence=float(result["confidence"]),
        strength=float(result["strength"]),
        cached=cached,
        comment_time=comment_time,
    )
    return {
        "record_id": stored["id"],
        "text": text,
        "clean_text": result["clean_text"],
        "label": int(result["label"]),
        "sentiment": str(result["sentiment"]),
        "confidence": float(result["confidence"]),
        "strength": float(result["strength"]),
        "cached": cached,
        "product_id": product_id,
    }


@app.get("/api/health")
def health() -> Dict[str, Any]:
    return api_response(
        {
            "status": "ok",
            "version": settings.app_version,
            "database_backend": settings.database_backend,
            "cache": prediction_cache.stats(),
        },
        message="服务正常",
    )


@app.post("/api/sentiment/single")
def analyze_single(payload: SingleSentimentRequest) -> Dict[str, Any]:
    try:
        result = _analyze_and_store(
            payload.text,
            product_id=payload.product_id or "",
            comment_time=payload.comment_time or "",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return api_response(result, message="分析成功")


def _read_upload(file: UploadFile) -> pd.DataFrame:
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    filename = file.filename or ""
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    try:
        if suffix in {"xlsx", "xls"}:
            return pd.read_excel(io.BytesIO(content))
        for encoding in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return pd.read_csv(io.BytesIO(content), encoding=encoding)
            except UnicodeDecodeError:
                continue
        return pd.read_csv(io.BytesIO(content), encoding="gbk", encoding_errors="replace")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {exc}") from exc


def _pick_text_column(df: pd.DataFrame) -> str:
    candidates = ["content", "comment", "text", "raw_text", "评论", "评价", "评论内容"]
    lower_map = {str(col).lower(): col for col in df.columns}
    for item in candidates:
        if item.lower() in lower_map:
            return lower_map[item.lower()]
    if len(df.columns) == 0:
        raise HTTPException(status_code=400, detail="文件没有可读取的列")
    return str(df.columns[0])


@app.post("/api/sentiment/batch")
def analyze_batch(
    file: UploadFile = File(...),
    product_id: str = Form(default=""),
) -> Dict[str, Any]:
    df = _read_upload(file)
    if len(df) > settings.max_upload_rows:
        raise HTTPException(status_code=400, detail=f"单次最多支持 {settings.max_upload_rows} 行")

    text_column = _pick_text_column(df)
    task_id = f"batch_{uuid.uuid4().hex[:12]}"
    create_batch_task(task_id, file.filename or "upload", len(df))

    results: List[Dict[str, Any]] = []
    failed_count = 0
    positive_count = 0
    negative_count = 0
    strength_sum = 0.0

    for idx, row in df.iterrows():
        row_product_id = str(row.get("product_id") or row.get("cat") or product_id or "")
        comment_time = str(row.get("comment_time") or row.get("time") or "")
        text = "" if pd.isna(row.get(text_column)) else str(row.get(text_column)).strip()
        if not text:
            failed_count += 1
            results.append({"row_index": int(idx), "success": False, "error": "评论文本为空"})
            continue
        try:
            item = _analyze_and_store(text, product_id=row_product_id, task_id=task_id, comment_time=comment_time)
            item["row_index"] = int(idx)
            item["success"] = True
            results.append(item)
            if item["sentiment"] == "positive":
                positive_count += 1
            else:
                negative_count += 1
            strength_sum += float(item["strength"])
        except Exception as exc:
            failed_count += 1
            results.append({"row_index": int(idx), "success": False, "error": str(exc)})

    success_count = len(results) - failed_count
    avg_strength = round(strength_sum / success_count, 2) if success_count else 0.0
    status = "completed" if failed_count == 0 else "completed_with_errors"
    update_error = "" if failed_count == 0 else f"{failed_count} rows failed"
    from .database import update_batch_task

    update_batch_task(
        task_id,
        status=status,
        success_count=success_count,
        failed_count=failed_count,
        positive_count=positive_count,
        negative_count=negative_count,
        avg_strength=avg_strength,
        error_message=update_error,
    )
    return api_response(
        {
            "task_id": task_id,
            "total": len(df),
            "success_count": success_count,
            "failed_count": failed_count,
            "positive_count": positive_count,
            "negative_count": negative_count,
            "avg_strength": avg_strength,
            "results": results,
        },
        message="批量分析完成",
    )


@app.get("/api/tasks/{task_id}")
def get_task(task_id: str) -> Dict[str, Any]:
    task = get_batch_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="批量任务不存在")
    return api_response(task)


@app.get("/api/history")
def history(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    sentiment: str = Query(default=""),
    product_id: str = Query(default=""),
    start_time: str = Query(default=""),
    end_time: str = Query(default=""),
) -> Dict[str, Any]:
    if sentiment and sentiment not in {"positive", "negative"}:
        raise HTTPException(status_code=400, detail="sentiment 只能是 positive 或 negative")
    return api_response(
        list_history(
            page,
            page_size,
            sentiment=sentiment,
            product_id=product_id,
            start_time=start_time,
            end_time=end_time,
        )
    )


@app.get("/api/statistics/summary")
def statistics_summary(product_id: str = Query(default="")) -> Dict[str, Any]:
    data = summary_statistics(product_id=product_id)
    data["cache"] = prediction_cache.stats()
    return api_response(data)


@app.get("/api/insights/keywords")
def insight_keywords(
    product_id: str = Query(default=""),
    top_k: int = Query(default=30, ge=1, le=100),
    start_time: str = Query(default=""),
    end_time: str = Query(default=""),
) -> Dict[str, Any]:
    records = analysis_records_for_insights(
        product_id=product_id,
        start_time=start_time,
        end_time=end_time,
    )
    keyword_data = extract_keywords(records, top_k=top_k)
    return api_response(
        {
            "product_id": product_id,
            "top_k": top_k,
            **keyword_data,
        },
        message="关键词统计成功",
    )


@app.get("/api/insights/aspects")
def insight_aspects(
    product_id: str = Query(default=""),
    start_time: str = Query(default=""),
    end_time: str = Query(default=""),
) -> Dict[str, Any]:
    records = analysis_records_for_insights(
        product_id=product_id,
        start_time=start_time,
        end_time=end_time,
    )
    aspect_data = score_aspects(records)
    return api_response(
        {
            "product_id": product_id,
            **aspect_data,
        },
        message="维度评分成功",
    )


@app.get("/api/summary/product")
def product_summary(
    product_id: str = Query(default=""),
    top_k: int = Query(default=10, ge=1, le=50),
    start_time: str = Query(default=""),
    end_time: str = Query(default=""),
) -> Dict[str, Any]:
    records = analysis_records_for_insights(
        product_id=product_id,
        start_time=start_time,
        end_time=end_time,
    )
    keyword_data = extract_keywords(records, top_k=top_k)
    aspect_data = score_aspects(records)
    summary = generate_rule_summary(keyword_data, aspect_data)
    return api_response(
        {
            "product_id": product_id,
            "record_count": len(records),
            **summary,
        },
        message="总结生成成功",
    )


@app.get("/api/export/{task_id}")
def export_task(task_id: str, file_type: str = Query(default="csv", pattern="^(csv|xlsx)$")) -> StreamingResponse:
    task = get_batch_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="批量任务不存在")
    records = records_by_task(task_id)
    if not records:
        raise HTTPException(status_code=404, detail="该任务没有可导出的记录")

    export_rows = _format_export_records(records)
    report_name = f"sentiment_report_{task_id}"

    if file_type == "xlsx":
        output = io.BytesIO()
        dataframe = pd.DataFrame(export_rows)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            dataframe.to_excel(writer, index=False, sheet_name="情感分析报告")
            worksheet = writer.sheets["情感分析报告"]
            _autosize_excel_columns(worksheet, dataframe)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": _content_disposition(f"{report_name}.xlsx")},
        )

    output = io.StringIO(newline="")
    output.write("\ufeff")
    fieldnames = [header for _, header in EXPORT_COLUMNS]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(export_rows)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": _content_disposition(f"{report_name}.csv")},
    )
